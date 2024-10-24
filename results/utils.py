from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.workbook import Workbook


def change_color_of_columns(filename: str, color: PatternFill):
    # Открываем созданный файл для добавления стилей
    workbook = load_workbook(filename=filename)
    worksheet = workbook.active


    # Применяем цвет к заголовкам (первая строка)
    for cell in worksheet[1]:  # Заголовки находятся в первой строке
        cell.fill = color

    format_columns(worksheet=worksheet)

    # Сохраняем изменения
    workbook.save(filename)
    print(f"Файл '{filename}' успешно создан с желтыми заголовками.")


def format_columns(worksheet: Workbook.active):
    # Автоматическая настройка ширины колонок
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Получаем букву колонки

        # Это ячейка конкретной колонки
        for cell in column:
            if cell.col_idx == 1 and cell.coordinate > "A1":  # Все ячейки ПЕРВОЙ колонки, кроме первой делаем типом "Ссылка"
                cell.style = 'Hyperlink'
            # А это значение конкретной колонки
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))  # Определяем максимальную длину содержимого

        adjusted_width = (max_length + 2)  # Добавляем небольшой отступ
        worksheet.column_dimensions[column_letter].width = adjusted_width  # Устанавливаем ширину колонки
